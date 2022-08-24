from pyModbusTCP.server import ModbusServer, DataBank
#import pose
from time import sleep
import openpyxl

#file = openpyxl.load_workbook('num_detected.xlsx')
#sheet = file.active

server = ModbusServer("192.168.137.1", 502, no_block=True)
server.start()

#destintion= {29: [34251, 26802, 33000], 30: [33646, 26802, 33000], 31: [33044, 26811, 33000], 32: [34242, 26441, 33000], 33:  [33639, 26460, 33000], 34:[33028, 26460, 33000]}

#def get_coordinate():
    #"""Calling and executing the camera's image processing script"""
    #return get_coordinates_copy.get_coordinates()


#def modbus_transfer(pos, sheet, act):
#    """send the start and end point coordinates from Excel file to the UR controller and also the acknowledgement bit
#    takes:
#    pose: which row of excel file to write
#    sheet: excel sheet
#    act: Gripping or releasing
#    """
#    if act == 0:
#        x1 = round(float(sheet.cell(row=pos, column=2).value))
#        y1 = round(float(sheet.cell(row=pos, column=3).value))
#        z1 = round(float(sheet.cell(row=pos, column=4).value))
#        DataBank.set_words(128, [int(x1)])
#        DataBank.set_words(129, [int(y1)])
#        DataBank.set_words(130, [int(z1)])
#    elif act == 1:
#        x2 = round(float(sheet.cell(row=pos, column=5).value))
#        y2 = round(float(sheet.cell(row=pos, column=6).value))
#        z2 = round(float(sheet.cell(row=pos, column=7).value))
#        DataBank.set_words(131, [int(x2)])
#        DataBank.set_words(132, [int(y2)])
#        DataBank.set_words(133, [int(z2)])
#   # Execution of grabbing the component
#    DataBank.set_words(134, [int(act)])
#    DataBank.set_bits(201, [True])
#    sleep(0.5)
#    DataBank.set_bits(201, [False])



def work():
    process = False
    try:
        #result = get_coordinate()
        if server.is_run:
            process = True

        while process:
            Trig1 = DataBank.get_bits(231)
            # Wait and check the trigger[start of the robot program] until true
            while Trig1 == [False]:
                print('Waiting')
                Trig1 = DataBank.get_bits(231)
            if Trig1 == [True]:
                print('robot started')

            #for i in result.keys():
            for i in range(0,2):
                Trig2 = DataBank.get_bits(232)
                while Trig2 == [False]:
                    sleep(0.01)
                    print('waiting for home position')
                    Trig2 = DataBank.get_bits(232)
                if Trig2 == [True]:
                    print('robot is homed')

                DataBank.set_words(138, [29808])
                DataBank.set_words(139, [26058])
                DataBank.set_words(140, [35500])
                DataBank.set_words(141, [0])
                DataBank.set_bits(142, [True])
                sleep(0.5)
                DataBank.set_bits(142, [False])

                #modbus_transfer(pos=components + 2, sheet=sheet, act=0)
                Trig3 = DataBank.get_bits(233)
                while Trig3 == [False]:
                    sleep(0.01)
                    print('Waiting for the robot to grab component number')
                    Trig3 = DataBank.get_bits(233)
                print('Component number Grabbed')
                """if Trig3 == [True]:
                    #pos1 = sheet.cell(row=components + 2, column=1).value
                    #modbus_transfer(pos=pos1, sheet=sheet, act=1)
                    DataBank.set_words(131, [destintion[keys][0]])
                    DataBank.set_words(132, [destintion[keys][1]])
                    DataBank.set_words(133, [destintion[keys][2]])
                    DataBank.set_words(134, [1])
                    DataBank.set_bits(201, [True])
                    sleep(0.5)
                    DataBank.set_bits(201, [False])

                    Trig4 = DataBank.get_bits(224)
                    while Trig4 == [False]:
                        sleep(0.01)
                        print(f'Waiting for the robot to insert component number {keys}')
                        Trig4 = DataBank.get_bits(224)
                    print(f'component number {keys} placed in slot')"""
            DataBank.set_bits(201, [True])
            break
    finally:
        #file.save('num_detected.xlsx')
        #file.close()
        server.stop()


if __name__ == '__main__':
    work()